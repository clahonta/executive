from pathlib import Path
import openpyxl



def load_excel_files(folder):

    context = []

    folder = Path(folder)

    if not folder.exists():
        return context


    for file in folder.glob("*.xlsx"):

        workbook = openpyxl.load_workbook(
            file,
            data_only=True,
            read_only=True  
        )

        workbook_text = []

        workbook_text.append(
            f"EXCEL FILE: {file.name}"
        )


        for sheet_name in workbook.sheetnames:

            sheet = workbook[sheet_name]

            workbook_text.append(
                f"\nSHEET: {sheet_name}"
            )


            for row in sheet.iter_rows(values_only=True):

                values = [
                    str(cell)
                    for cell in row
                    if cell is not None
                ]

                if values:
                    workbook_text.append(
                        " | ".join(values)
                    )


        context.append(
            "\n".join(workbook_text)
        )


    return context
